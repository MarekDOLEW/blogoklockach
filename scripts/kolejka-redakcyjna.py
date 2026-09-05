import json, datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

kat = json.load(open('/home/user/blogoklockach/src/data/katalog.json'))
sety = json.load(open('/home/user/blogoklockach/src/data/sety.json'))
opisane = set(sety.keys())

rows = [(s, r) for s, v in kat.items() if s != '_meta' and isinstance(v, list) for r in v]
kand = [(s, r) for s, r in rows
        if isinstance(r.get('rok'), int) and 2020 <= r['rok'] <= 2026
        and r.get('status') == 'dostepny'
        and r.get('cena_katalogowa') not in (None, '', 0)
        and str(r.get('numer')) not in opisane]

# dedup po numerze — zostawiamy rekord z nowszym rocznikiem
best = {}
for s, r in kand:
    n = str(r['numer'])
    if n not in best or r['rok'] > best[n][1]['rok']:
        best[n] = (s, r)
data = sorted(best.values(), key=lambda x: (-float(x[1]['cena_katalogowa']), x[1]['rok'], str(x[1]['numer'])))

ARIAL = 'Arial'
wb = Workbook()
ws = wb.active
ws.title = 'Kolejka redakcyjna'

hdr_fill = PatternFill('solid', fgColor='1F3864')
hdr_font = Font(name=ARIAL, size=11, bold=True, color='FFFFFF')
thin = Side(style='thin', color='BFBFBF')
border = Border(bottom=thin)

headers = ['Numer', 'Nazwa', 'Rok wydania', 'Seria', 'Cena katalogowa (zł)']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

band = PatternFill('solid', fgColor='F2F2F2')
for i, (seria, r) in enumerate(data, start=2):
    ws.cell(row=i, column=1, value=str(r['numer'])).alignment = Alignment(horizontal='left')
    ws.cell(row=i, column=2, value=r.get('nazwa'))
    ws.cell(row=i, column=3, value=r['rok']).alignment = Alignment(horizontal='center')
    ws.cell(row=i, column=4, value=seria)
    pc = ws.cell(row=i, column=5, value=float(r['cena_katalogowa']))
    pc.number_format = '#,##0.00'
    for c in range(1, 6):
        cl = ws.cell(row=i, column=c)
        cl.font = Font(name=ARIAL, size=10)
        cl.border = border
        if i % 2 == 0:
            cl.fill = band

widths = [10, 62, 13, 22, 20]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 26
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:E{len(data)+1}'

# --- arkusz informacyjny ---
ws2 = wb.create_sheet('Metodologia')
per_year = Counter(r['rok'] for _, r in data)
info = [
    ('KOLEJKA REDAKCYJNA — zestawy LEGO bez opisu na tylkoklocki.pl', ''),
    ('', ''),
    ('Data wygenerowania', datetime.date.today().isoformat()),
    ('Liczba pozycji', len(data)),
    ('', ''),
    ('KRYTERIA (wszystkie muszą być spełnione jednocześnie)', ''),
    ('1. Rocznik', '2020–2026'),
    ('2. Status w katalogu', 'dostepny (nadal w sprzedaży, nie EOL)'),
    ('3. Cena katalogowa', 'znana, w złotych'),
    ('4. Opis redakcyjny', 'BRAK — numer nieobecny w src/data/sety.json'),
    ('', ''),
    ('ŹRÓDŁA DANYCH', ''),
    ('Numer, nazwa PL, rok, seria, cena', 'src/data/katalog.json (aktualizacja 2026-09-02)'),
    ('Lista opisanych zestawów', 'src/data/sety.json (303 pozycje)'),
    ('', ''),
    ('SORTOWANIE', 'cena katalogowa malejąco — droższy zestaw = wyższa prowizja afiliacyjna'),
    ('', ''),
    ('ROZKŁAD PO ROCZNIKACH', ''),
]
for rok in sorted(per_year):
    info.append((str(rok), per_year[rok]))
info += [
    ('', ''),
    ('UWAGI', ''),
    ('Zestaw 40824 (Tweety)', 'w katalogu występuje dwukrotnie — jako Seasonal/2025 i Looney Tunes/2026. '
                              'W tabeli zostawiono nowszy wpis. Do zweryfikowania przy opisie.'),
    ('Historia liczby', '04.09: 809 pozycji (start kolejki). 05.09: 721 — sesja redakcyjna zamknęła 88 '
                        'najdroższych, doszły pojedyncze nowe zestawy z katalogu.'),
    ('Poza kolejką', 'zestawy z ceną w zł, ale ze statusem EOL — nie konwertują poza rynkiem wtórnym, '
                     'więc nie wchodzą do kolejki.'),
    ('Plik odświeżany', 'przez runnera Scout Nowości przy każdym przebiegu — liczba maleje w miarę pracy '
                        'sesji redakcyjnej.'),
]
for i, (a, b) in enumerate(info, start=1):
    ca = ws2.cell(row=i, column=1, value=a)
    cb = ws2.cell(row=i, column=2, value=b)
    ca.font = Font(name=ARIAL, size=10, bold=(b == '' and a != ''))
    cb.font = Font(name=ARIAL, size=10)
    cb.alignment = Alignment(wrap_text=True, vertical='top')
ws2.cell(row=1, column=1).font = Font(name=ARIAL, size=13, bold=True)
ws2.column_dimensions['A'].width = 42
ws2.column_dimensions['B'].width = 78

out = '/home/user/blogoklockach/materialy/kolejka-redakcyjna.xlsx'
wb.save(out)
print('zapisano', out, 'pozycji:', len(data))
print('po rocznikach:', sorted(per_year.items()))
